"""
Genera el archivo de datos editable (datos_pentur.xlsx) que alimenta el dashboard.
Estructura TIDY: una fila por (horizonte, nivel/pilar, línea de acción).
El contenido es editable desde Excel; la ESTRUCTURA del mapa (perspectivas y
niveles) es fija y se define en config_mapa.py.

Ejecutar:  python generar_datos.py
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Claves de perspectiva y nivel/pilar deben coincidir EXACTAMENTE con config_mapa.py
DC = "Desarrollo de capacidades y condiciones"
GT = "Gestión territorial"
GV_VIS = "Gestión de visitantes"
GV = "Generación de valor e innovación"

# ------------------------------------------------------------------ #
#  DATOS  (fuente: Formulación del plan estratégico PENTUR 2036, IDOM)
#  cols: horizonte, perspectiva, nivel_pilar, linea_accion, indicador,
#        meta_emergente, meta_ascenso, meta_referente
# ------------------------------------------------------------------ #
rows = []

def add(horizonte, perspectiva, nivel, linea, indicador, m_em, m_as, m_re):
    rows.append(dict(horizonte=horizonte, perspectiva=perspectiva, nivel_pilar=nivel,
                     linea_accion=linea, indicador=indicador,
                     meta_emergente=m_em, meta_ascenso=m_as, meta_referente=m_re))

# ============================ ACTIVACIÓN 2027-2029 ============================
H = "Activación (2027-2029)"
add(H, DC, "Sistema de inteligencia turística",
    "Construir repositorio de inteligencia de mercado accesible para el sector público y privado",
    "N.º de módulos del repositorio de inteligencia turística operativos y accesibles",
    "1 módulo (Nacional)", "2 módulos (Nacional)", "3 módulos (Nacional)")
add(H, DC, "Sistema de inteligencia turística",
    "Elaborar estudios de prospectiva y tendencias de la demanda turística por segmento y mercado emisor",
    "N.º de estudios de prospectiva de la demanda turística publicados",
    "1 estudio", "2 estudios", "3 estudios")
add(H, DC, "Sistema de inteligencia turística",
    "Publicar análisis de equilibrio entre oferta y demanda turística por destino para orientar la inversión",
    "N.º de análisis de equilibrio de oferta y demanda por destino publicados",
    "1 análisis", "2 análisis", "3 análisis")
add(H, DC, "Marco normativo sectorial",
    "Actualizar el marco normativo del sector turismo, incluyendo la creación de la figura del gestor de experiencias turísticas",
    "N.º de disposiciones del marco normativo turístico actualizadas o creadas",
    "1 disposición (figura del gestor y reglamento)", "2 disposiciones", "3 disposiciones")
# Gestión territorial - Oferta
add(H, GT, "Oferta",
    "Diseñar productos y experiencias turísticas con base en estudios de mercado y tendencias",
    "N.º de productos turísticos diseñados y operativos en el destino",
    "1 producto", "2 productos", "4 productos")
add(H, GT, "Oferta",
    "Brindar asistencia técnica para la formulación de planes de negocio a empresas turísticas",
    "N.º de empresas turísticas con plan de negocio formulado",
    "10 empresas", "20 empresas", "30 empresas")
add(H, GT, "Oferta",
    "Vincular empresas turísticas con fondos concursables (Turismo Emprende, ProInnóvate, ProCompite)",
    "N.º de empresas turísticas postulantes a fondos concursables del sector",
    "5 empresas", "15 empresas", "30 empresas")
add(H, GT, "Oferta",
    "Ampliar la cobertura y los recursos de fondos concursables desde MINCETUR",
    "% de regiones con acceso activo a fondos concursables del sector turismo",
    "60% (Nacional)", "65% (Nacional)", "70% (Nacional)")
add(H, GT, "Oferta",
    "Fortalecer el talento humano en la prestación de servicios turísticos (CENFOTUR)",
    "N.º de personas capacitadas en prestación de servicios turísticos",
    "50 personas", "100 personas", "200 personas")
# Promoción y demanda
add(H, GT, "Promoción y demanda",
    "Diseñar estrategias de promoción diferenciadas por público objetivo y mercado emisor, con base en investigación de mercados",
    "N.º de estrategias de promoción turística diferenciadas por mercado emisor implementadas",
    "1 estrategia", "3 estrategias", "5 estrategias")
add(H, GT, "Promoción y demanda",
    "Fortalecer la marca Perú y las marcas de los destinos priorizados",
    "N.º de destinos de la región con marca turística oficial activa en canales de distribución",
    "1 destino", "2 destinos", "Todos los destinos")
add(H, GT, "Promoción y demanda",
    "Consolidar la participación en ferias y eventos de promoción turística (Peru Travel, LBIT)",
    "N.º de ferias y eventos de promoción turística en los que participa el destino",
    "1 feria", "2 ferias", "3 ferias")
# Infraestructura
add(H, GT, "Infraestructura",
    "Poner en valor atractivos turísticos, centros soporte y espacios urbanos vinculados al destino",
    "N.º de atractivos turísticos y centros soporte con puesta en valor ejecutada",
    "1 atractivo o centro soporte", "2 atractivos y 1 centro soporte", "4 atractivos y 1 centro soporte")
add(H, GT, "Infraestructura",
    "Implementar sistemas señaléticos integrales con enfoque de destino (urbana, vial y de atractivo)",
    "% de atractivos y puntos de acceso del destino con señalética turística instalada",
    "30%", "50%", "70%")
add(H, GT, "Infraestructura",
    "Promover la inversión privada en planta turística (hotelería, restauración y servicios conexos)",
    "N.º de establecimientos de planta turística privada habilitados en el destino",
    "2 establecimientos", "5 establecimientos", "10 establecimientos")
add(H, GT, "Infraestructura",
    "Incorporar criterios de accesibilidad universal en todos los proyectos de inversión pública turística",
    "% de proyectos de inversión pública turística con criterios de accesibilidad universal incorporados",
    "30%", "60%", "100%")
# Gobernanza
add(H, GT, "Gobernanza",
    "Crear, fortalecer o consolidar entes gestores de destino según estado institucional de cada uno",
    "N.º de entes gestores de destino turístico en funcionamiento activo y verificable",
    "1 (en creación)", "1 (operativo)", "1 (consolidado)")
add(H, GT, "Gobernanza",
    "Asignar, incrementar y garantizar presupuesto para la función turismo en gobiernos regionales y municipales",
    "% de ejecución del presupuesto regional asignado a la función turismo",
    "60%", "75%", "90%")
add(H, GT, "Gobernanza",
    "Fortalecer gremios turísticos privados con representatividad mínima verificable",
    "% del sector privado turístico representado en el gremio regional reconocido por MINCETUR",
    "20%", "30%", "50%")
add(H, GT, "Gobernanza",
    "Capacitar a DIRCETURES y GERCETURES en formulación, gestión y ejecución de proyectos turísticos",
    "% de funcionarios de la DIRCETUR capacitados en formulación y gestión de proyectos",
    "50%", "75%", "100%")
add(H, GT, "Gobernanza",
    "Articular mecanismos de financiamiento internacional para el sector turismo",
    "N.º de proyectos de cooperación internacional para el sector turismo en gestión o ejecución",
    "1 proyecto", "1 proyecto", "1 proyecto")
add(H, GT, "Gobernanza",
    "Desarrollar el capital humano en gestión institucional turística",
    "N.º de funcionarios públicos con capacidades de gestión institucional fortalecidas",
    "3 funcionarios", "8 funcionarios", "15 funcionarios")
# Gestión de visitantes
add(H, GV_VIS, "Calidad",
    "Implementar el sistema CALTUR en empresas, personas, sitios y destinos",
    "N.º de prestadores de servicios turísticos con certificación CALTUR vigente",
    "5 prestadores", "15 prestadores", "30 prestadores")
add(H, GV_VIS, "Seguridad",
    "Desarrollar protocolos de seguridad ciudadana, operativa, sanitaria y de gestión de riesgos en los destinos",
    "% de destinos priorizados de la región con protocolo integral de seguridad turística implementado",
    "10%", "50%", "100%")
add(H, GV_VIS, "Inclusión",
    "Incorporar criterios de accesibilidad universal en infraestructura, oferta y gobernanza de los destinos",
    "N.º de atractivos y servicios turísticos con condiciones de accesibilidad universal",
    "1 atractivo", "3 atractivos", "6 atractivos")
add(H, GV_VIS, "Sostenibilidad",
    "Promover modelos de operación turística que optimicen el uso de recursos y reduzcan impactos negativos",
    "N.º de operadores turísticos con prácticas de sostenibilidad implementadas",
    "5 operadores", "15 operadores", "30 operadores")
# Generación de valor
add(H, GV, "Competitividad turística",
    "Posicionar a Perú en el ranking de competitividad turística regional",
    "Posición de Perú en el WEF Travel & Tourism Development Index (TTDI)",
    "Mantener posición 2025 (Nacional)", "Mantener posición 2025 (Nacional)", "Mantener posición 2025 (Nacional)")
add(H, GV, "Arribos y viajes",
    "Incrementar los arribos internacionales y los viajes nacionales hacia los destinos priorizados",
    "Tasa de variación del número de visitantes al destino respecto al año base (2025)",
    "+10%", "+20%", "+30%")
add(H, GV, "Inversión pública",
    "Incrementar la inversión pública ejecutada en turismo a nivel nacional y regional",
    "% de ejecución de la inversión pública nacional destinada al sector turismo",
    "70% (Nacional)", "70% (Nacional)", "70% (Nacional)")
add(H, GV, "Aporte al PIB",
    "Aumentar la contribución del turismo al PIB nacional",
    "Participación del turismo en el PBI nacional (puntos porcentuales)",
    "Igual al año base (Nacional)", "Igual al año base (Nacional)", "Igual al año base (Nacional)")

# ========================= FORTALECIMIENTO 2030-2032 =========================
H = "Fortalecimiento (2030-2032)"
add(H, DC, "Sistema de inteligencia turística",
    "Construir repositorio de inteligencia de mercado accesible para el sector público y privado",
    "N.º de módulos del repositorio de inteligencia turística operativos y accesibles",
    "2 módulos (Nacional)", "2 módulos (Nacional)", "2 módulos (Nacional)")
add(H, DC, "Sistema de inteligencia turística",
    "Elaborar estudios de prospectiva y tendencias de la demanda turística por segmento y mercado emisor",
    "N.º de estudios de prospectiva de la demanda turística publicados",
    "2 estudios", "3 estudios", "4 estudios")
add(H, DC, "Sistema de inteligencia turística",
    "Publicar análisis de equilibrio entre oferta y demanda turística por destino para orientar la inversión",
    "N.º de análisis de equilibrio de oferta y demanda por destino publicados",
    "2 análisis", "3 análisis", "4 análisis")
add(H, DC, "Marco normativo sectorial",
    "Actualizar el marco normativo del sector turismo, incluyendo la creación de la figura del gestor de experiencias turísticas",
    "N.º de disposiciones del marco normativo turístico actualizadas o creadas",
    "3 disposiciones (Nacional)", "3 disposiciones (Nacional)", "3 disposiciones (Nacional)")
add(H, GT, "Oferta",
    "Diseñar productos y experiencias turísticas con base en estudios de mercado y tendencias",
    "N.º de productos turísticos diseñados y operativos en el destino",
    "2 productos", "4 productos", "6 productos")
add(H, GT, "Oferta",
    "Brindar asistencia técnica para la formulación de planes de negocio a empresas turísticas",
    "N.º de empresas turísticas con plan de negocio formulado",
    "20 empresas", "35 empresas", "50 empresas")
add(H, GT, "Oferta",
    "Vincular empresas turísticas con fondos concursables (Turismo Emprende, ProInnóvate, ProCompite)",
    "N.º de empresas turísticas postulantes a fondos concursables del sector",
    "15 empresas", "25 empresas", "45 empresas")
add(H, GT, "Oferta",
    "Ampliar la cobertura y los recursos de fondos concursables desde MINCETUR",
    "% de regiones con acceso activo a fondos concursables del sector turismo",
    "80% (Nacional)", "70% (Nacional)", "60% (Nacional)")
add(H, GT, "Oferta",
    "Fortalecer el talento humano en la prestación de servicios turísticos (CENFOTUR)",
    "N.º de personas capacitadas en prestación de servicios turísticos",
    "100 personas", "200 personas", "350 personas")
add(H, GT, "Promoción y demanda",
    "Diseñar estrategias de promoción diferenciadas por público objetivo y mercado emisor, con base en investigación de mercados",
    "N.º de estrategias de promoción turística diferenciadas por mercado emisor implementadas",
    "1 estrategia", "2 estrategias", "3 estrategias")
add(H, GT, "Promoción y demanda",
    "Fortalecer la marca Perú y las marcas de los destinos priorizados",
    "N.º de destinos de la región con marca turística oficial activa en canales de distribución",
    "1 destino", "1 destino", "2 destinos")
add(H, GT, "Promoción y demanda",
    "Consolidar la participación en ferias y eventos de promoción turística (Peru Travel, LBIT)",
    "N.º de ferias y eventos de promoción turística en los que participa el destino",
    "2 ferias", "2 ferias", "3 ferias")
add(H, GT, "Infraestructura",
    "Poner en valor atractivos turísticos, centros soporte y espacios urbanos vinculados al destino",
    "N.º de atractivos turísticos y centros soporte con puesta en valor ejecutada",
    "2 atractivos", "4 atractivos", "6 atractivos")
add(H, GT, "Infraestructura",
    "Implementar sistemas señaléticos integrales con enfoque de destino (urbana, vial y de atractivo)",
    "% de atractivos y puntos de acceso del destino con señalética turística instalada",
    "60%", "80%", "100%")
add(H, GT, "Infraestructura",
    "Promover la inversión privada en planta turística (hotelería, restauración y servicios conexos)",
    "N.º de establecimientos de planta turística privada habilitados en el destino",
    "5 establecimientos", "10 establecimientos", "20 establecimientos")
add(H, GT, "Infraestructura",
    "Incorporar criterios de accesibilidad universal en todos los proyectos de inversión pública turística",
    "% de proyectos de inversión pública turística con criterios de accesibilidad universal incorporados",
    "60%", "80%", "100%")
add(H, GT, "Gobernanza",
    "Crear, fortalecer o consolidar entes gestores de destino según estado institucional de cada uno",
    "N.º de entes gestores de destino turístico en funcionamiento activo y verificable",
    "1 (fortalecido)", "1 (potenciado)", "2 (consolidados)")
add(H, GT, "Gobernanza",
    "Asignar, incrementar y garantizar presupuesto para la función turismo en gobiernos regionales y municipales",
    "% de ejecución del presupuesto regional asignado a la función turismo",
    "75%", "85%", "95%")
add(H, GT, "Gobernanza",
    "Fortalecer gremios turísticos privados con representatividad mínima verificable",
    "% del sector privado turístico representado en el gremio regional reconocido por MINCETUR",
    "30%", "40%", "60%")
add(H, GT, "Gobernanza",
    "Capacitar a DIRCETURES y GERCETURES en formulación, gestión y ejecución de proyectos turísticos",
    "% de funcionarios de la DIRCETUR capacitados en formulación y gestión de proyectos",
    "75%", "90%", "100%")
add(H, GT, "Gobernanza",
    "Articular mecanismos de financiamiento internacional para el sector turismo",
    "N.º de proyectos de cooperación internacional para el sector turismo en gestión o ejecución",
    "1 proyecto", "1 proyecto", "2 proyectos")
add(H, GT, "Gobernanza",
    "Desarrollar el capital humano en gestión institucional turística",
    "N.º de funcionarios públicos con capacidades de gestión institucional fortalecidas",
    "8 funcionarios", "15 funcionarios", "25 funcionarios")
add(H, GV_VIS, "Calidad",
    "Implementar el sistema CALTUR en empresas, personas, sitios y destinos",
    "N.º de prestadores de servicios turísticos con certificación CALTUR vigente",
    "15 prestadores", "30 prestadores", "50 prestadores")
add(H, GV_VIS, "Seguridad",
    "Desarrollar protocolos de seguridad ciudadana, operativa, sanitaria y de gestión de riesgos en los destinos",
    "% de destinos priorizados de la región con protocolo integral de seguridad turística implementado",
    "50%", "100%", "100%")
add(H, GV_VIS, "Inclusión",
    "Incorporar criterios de accesibilidad universal en infraestructura, oferta y gobernanza de los destinos",
    "N.º de atractivos y servicios turísticos con condiciones de accesibilidad universal",
    "3 atractivos", "6 atractivos", "10 atractivos")
add(H, GV_VIS, "Sostenibilidad",
    "Promover modelos de operación turística que optimicen el uso de recursos y reduzcan impactos negativos",
    "N.º de operadores turísticos con prácticas de sostenibilidad implementadas",
    "20 operadores", "35 operadores", "60 operadores")
add(H, GV, "Competitividad turística",
    "Posicionar a Perú en el ranking de competitividad turística regional",
    "Posición de Perú en el WEF Travel & Tourism Development Index (TTDI)",
    "Avanzar 2 posiciones respecto a 2025 (Nacional)", "Avanzar 2 posiciones respecto a 2025 (Nacional)", "Avanzar 2 posiciones respecto a 2025 (Nacional)")
add(H, GV, "Arribos y viajes",
    "Incrementar los arribos internacionales y los viajes nacionales hacia los destinos priorizados",
    "Tasa de variación del número de visitantes al destino respecto al año base (2025)",
    "+30%", "+50%", "+70%")
add(H, GV, "Inversión pública",
    "Incrementar la inversión pública ejecutada en turismo a nivel nacional y regional",
    "% de ejecución de la inversión pública nacional destinada al sector turismo",
    "85% (Nacional)", "85% (Nacional)", "85% (Nacional)")
add(H, GV, "Aporte al PIB",
    "Aumentar la contribución del turismo al PIB nacional",
    "Participación del turismo en el PBI nacional (puntos porcentuales)",
    "+0.3 pp respecto al año base (Nacional)", "+0.3 pp respecto al año base (Nacional)", "+0.3 pp respecto al año base (Nacional)")

# =========================== ESCALAMIENTO 2033-2036 ==========================
H = "Escalamiento (2033-2036)"
add(H, DC, "Sistema de inteligencia turística",
    "Construir repositorio de inteligencia de mercado accesible para el sector público y privado",
    "N.º de módulos del repositorio de inteligencia turística operativos y accesibles",
    "3 módulos (Nacional)", "4 módulos (Nacional)", "5 módulos (Nacional)")
add(H, DC, "Sistema de inteligencia turística",
    "Elaborar estudios de prospectiva y tendencias de la demanda turística por segmento y mercado emisor",
    "N.º de estudios de prospectiva de la demanda turística publicados",
    "3 estudios", "4 estudios", "5 estudios")
add(H, DC, "Sistema de inteligencia turística",
    "Publicar análisis de equilibrio entre oferta y demanda turística por destino para orientar la inversión",
    "N.º de análisis de equilibrio de oferta y demanda por destino publicados",
    "3 análisis", "4 análisis", "5 análisis")
add(H, DC, "Marco normativo sectorial",
    "Actualizar el marco normativo del sector turismo, incluyendo la creación de la figura del gestor de experiencias turísticas",
    "N.º de disposiciones del marco normativo turístico actualizadas o creadas",
    "5 disposiciones (Nacional)", "5 disposiciones (Nacional)", "5 disposiciones (Nacional)")
add(H, GT, "Oferta",
    "Diseñar productos y experiencias turísticas con base en estudios de mercado y tendencias",
    "N.º de productos turísticos diseñados y operativos en el destino",
    "4 productos", "6 productos", "10 productos")
add(H, GT, "Oferta",
    "Brindar asistencia técnica para la formulación de planes de negocio a empresas turísticas",
    "N.º de empresas turísticas con plan de negocio formulado",
    "35 empresas", "55 empresas", "80 empresas")
add(H, GT, "Oferta",
    "Vincular empresas turísticas con fondos concursables (Turismo Emprende, ProInnóvate, ProCompite)",
    "N.º de empresas turísticas postulantes a fondos concursables del sector",
    "30 empresas", "50 empresas", "70 empresas")
add(H, GT, "Oferta",
    "Ampliar la cobertura y los recursos de fondos concursables desde MINCETUR",
    "% de regiones con acceso activo a fondos concursables del sector turismo",
    "100% (Nacional)", "100% (Nacional)", "100% (Nacional)")
add(H, GT, "Oferta",
    "Fortalecer el talento humano en la prestación de servicios turísticos (CENFOTUR)",
    "N.º de personas capacitadas en prestación de servicios turísticos",
    "200 personas", "350 personas", "500 personas")
add(H, GT, "Promoción y demanda",
    "Diseñar estrategias de promoción diferenciadas por público objetivo y mercado emisor, con base en investigación de mercados",
    "N.º de estrategias de promoción turística diferenciadas por mercado emisor implementadas",
    "2 estrategias", "3 estrategias", "4 estrategias")
add(H, GT, "Promoción y demanda",
    "Fortalecer la marca Perú y las marcas de los destinos priorizados",
    "N.º de destinos de la región con marca turística oficial activa en canales de distribución",
    "1 destino", "2 destinos", "3 destinos")
add(H, GT, "Promoción y demanda",
    "Consolidar la participación en ferias y eventos de promoción turística (Peru Travel, LBIT)",
    "N.º de ferias y eventos de promoción turística en los que participa el destino",
    "2 ferias", "3 ferias", "4 ferias")
add(H, GT, "Infraestructura",
    "Poner en valor atractivos turísticos, centros soporte y espacios urbanos vinculados al destino",
    "N.º de atractivos turísticos y centros soporte con puesta en valor ejecutada",
    "4 atractivos", "6 atractivos", "10 atractivos")
add(H, GT, "Infraestructura",
    "Implementar sistemas señaléticos integrales con enfoque de destino (urbana, vial y de atractivo)",
    "% de atractivos y puntos de acceso del destino con señalética turística instalada",
    "80%", "100%", "100%")
add(H, GT, "Infraestructura",
    "Promover la inversión privada en planta turística (hotelería, restauración y servicios conexos)",
    "N.º de establecimientos de planta turística privada habilitados en el destino",
    "10 establecimientos", "20 establecimientos", "40 establecimientos")
add(H, GT, "Infraestructura",
    "Incorporar criterios de accesibilidad universal en todos los proyectos de inversión pública turística",
    "% de proyectos de inversión pública turística con criterios de accesibilidad universal incorporados",
    "100%", "100%", "100%")
add(H, GT, "Gobernanza",
    "Crear, fortalecer o consolidar entes gestores de destino según estado institucional de cada uno",
    "N.º de entes gestores de destino turístico en funcionamiento activo y verificable",
    "1 (consolidado)", "2 (ampliados)", "3 (referentes)")
add(H, GT, "Gobernanza",
    "Asignar, incrementar y garantizar presupuesto para la función turismo en gobiernos regionales y municipales",
    "% de ejecución del presupuesto regional asignado a la función turismo",
    "85%", "95%", "98%")
add(H, GT, "Gobernanza",
    "Fortalecer gremios turísticos privados con representatividad mínima verificable",
    "% del sector privado turístico representado en el gremio regional reconocido por MINCETUR",
    "40%", "55%", "70%")
add(H, GT, "Gobernanza",
    "Capacitar a DIRCETURES y GERCETURES en formulación, gestión y ejecución de proyectos turísticos",
    "% de funcionarios de la DIRCETUR capacitados en formulación y gestión de proyectos",
    "100%", "100%", "100%")
add(H, GT, "Gobernanza",
    "Articular mecanismos de financiamiento internacional para el sector turismo",
    "N.º de proyectos de cooperación internacional para el sector turismo en gestión o ejecución",
    "1 proyecto", "2 proyectos", "2 proyectos")
add(H, GT, "Gobernanza",
    "Desarrollar el capital humano en gestión institucional turística",
    "N.º de funcionarios públicos con capacidades de gestión institucional fortalecidas",
    "15 funcionarios", "25 funcionarios", "40 funcionarios")
add(H, GV_VIS, "Calidad",
    "Implementar el sistema CALTUR en empresas, personas, sitios y destinos",
    "N.º de prestadores de servicios turísticos con certificación CALTUR vigente",
    "25 prestadores", "50 prestadores", "80 prestadores")
add(H, GV_VIS, "Seguridad",
    "Desarrollar protocolos de seguridad ciudadana, operativa, sanitaria y de gestión de riesgos en los destinos",
    "% de destinos priorizados de la región con protocolo integral de seguridad turística implementado",
    "100%", "100%", "100%")
add(H, GV_VIS, "Inclusión",
    "Incorporar criterios de accesibilidad universal en infraestructura, oferta y gobernanza de los destinos",
    "N.º de atractivos y servicios turísticos con condiciones de accesibilidad universal",
    "6 atractivos", "10 atractivos", "15 atractivos")
add(H, GV_VIS, "Sostenibilidad",
    "Promover modelos de operación turística que optimicen el uso de recursos y reduzcan impactos negativos",
    "N.º de operadores turísticos con prácticas de sostenibilidad implementadas",
    "40 operadores", "70 operadores", "120 operadores")
add(H, GV, "Competitividad turística",
    "Posicionar a Perú en el ranking de competitividad turística regional",
    "Posición de Perú en el WEF Travel & Tourism Development Index (TTDI)",
    "Avanzar 4 posiciones respecto a 2025 (Nacional)", "Avanzar 4 posiciones respecto a 2025 (Nacional)", "Figurar entre los 3 primeros destinos de América del Sur")
add(H, GV, "Arribos y viajes",
    "Incrementar los arribos internacionales y los viajes nacionales hacia los destinos priorizados",
    "Tasa de variación del número de visitantes al destino respecto al año base (2025)",
    "+60%", "+100%", "+175% (receptivo) / +100% (interno)")
add(H, GV, "Inversión pública",
    "Incrementar la inversión pública ejecutada en turismo a nivel nacional y regional",
    "% de ejecución de la inversión pública nacional destinada al sector turismo",
    "90% (Nacional)", "90% (Nacional)", "90% (Nacional)")
add(H, GV, "Aporte al PIB",
    "Aumentar la contribución del turismo al PIB nacional",
    "Participación del turismo en el PBI nacional (puntos porcentuales)",
    "+0.5 pp respecto al año base (Nacional)", "+0.5 pp respecto al año base (Nacional)", "+1.0 pp al cierre del plan")

# ------------------------------------------------------------------ #
df = pd.DataFrame(rows, columns=[
    "horizonte", "perspectiva", "nivel_pilar", "linea_accion", "indicador",
    "meta_emergente", "meta_ascenso", "meta_referente"])

OUT = "datos_pentur.xlsx"
with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
    df.to_excel(xl, sheet_name="lineas_de_accion", index=False)

# --- formato ligero + leyenda de edición --------------------------- #
wb = load_workbook(OUT)
ws = wb["lineas_de_accion"]
head_fill = PatternFill("solid", fgColor="C8102E")   # rojo Perú
head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
body_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
widths = {"A": 26, "B": 30, "C": 26, "D": 52, "E": 52, "F": 30, "G": 30, "H": 34}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for c in ws[1]:
    c.fill = head_fill; c.font = head_font
    c.alignment = Alignment(vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.font = body_font; c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# hoja de instrucciones
info = wb.create_sheet("LÉEME", 0)
info["A1"] = "Fuente de datos del dashboard PENTUR 2036"
info["A1"].font = Font(name="Arial", bold=True, size=13)
notas = [
    "",
    "Cómo actualizar el contenido del dashboard:",
    "• Edite ÚNICAMENTE la hoja 'lineas_de_accion'.",
    "• NO cambie los encabezados de columna ni los nombres de la columna 'perspectiva' / 'nivel_pilar'",
    "  (definen la estructura fija del mapa estratégico y del cubo).",
    "",
    "Valores admitidos (deben escribirse EXACTAMENTE así):",
    "  horizonte    → 'Activación (2027-2029)', 'Fortalecimiento (2030-2032)', 'Escalamiento (2033-2036)'",
    "  perspectiva  → 'Desarrollo de capacidades y condiciones', 'Gestión territorial',",
    "                 'Gestión de visitantes', 'Generación de valor e innovación'",
    "  nivel_pilar  → según la perspectiva (p. ej. 'Oferta', 'Infraestructura', 'Sistema de inteligencia turística'…)",
    "",
    "Columnas de contenido editables libremente:",
    "  linea_accion, indicador, meta_emergente, meta_ascenso, meta_referente",
    "",
    "Para AÑADIR una línea de acción: agregue una fila nueva respetando horizonte/perspectiva/nivel_pilar.",
    "Para AÑADIR un nuevo nivel/pilar al mapa: coordinar, requiere ajustar la estructura fija (config_mapa.py).",
    "",
    "Tras guardar, recargue el dashboard (botón 'Recargar datos' o refresque el navegador).",
]
for i, t in enumerate(notas, start=2):
    info[f"A{i}"] = t
    info[f"A{i}"].font = Font(name="Arial", size=10, bold=t.endswith(":"))
info.column_dimensions["A"].width = 110

wb.save(OUT)
print(f"OK -> {OUT}  ({len(df)} filas, {df['horizonte'].nunique()} horizontes)")
